#!/usr/bin/env python
# -*- coding=UTF-8 -*-
'''
@ Author: shy
@ Email: yushuibo@ebupt.com / hengchen2005@gmail.com
@ Version: v1.0
@ Description: -
@ Since: 2020/5/28 17:50
'''

from __future__ import print_function
from __future__ import unicode_literals

import os
import sys
from zipfile import BadZipfile

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle
from openpyxl.styles.named_styles import NamedStyleList
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

from utils import chinese_counter
from simplelogger import SimpleLogger

logger = SimpleLogger()


class ExcelUtil(object):

    def __init__(self):
        __style = NamedStyle(name='simple')
        __style.font = Font(bold=True, size=11, color=colors.BLACK)
        __style.fill = PatternFill(patternType='solid', fgColor='FABF8F')
        bd = Side(style='thin', color=colors.BLACK)
        self.__border = Border(left=bd, top=bd, right=bd, bottom=bd)
        __style.border = self.__border
        self.__alignment = Alignment(horizontal='left', vertical='center')
        __style.alignment = self.__alignment
        self.__header_style = __style

    def write(self, filename, data, sheet=None, sheet_index=0, has_title=False, is_overwrite=False):
        '''Write data to excel file.

        Arguments:
            file {str} -- The absolute name of excel file
            data {list} -- The data list which will write to excel file.

        Keyword Arguments:
            sheet {str} -- The name of destnation sheet.
            sheet_index {int} -- The index of sheet, start with 0.
            has_title {bool} -- True means the first row as table header (default: {False})
            is_overwrite {bool} -- When the destnation file exist, it will be overwrite
                when this flag is true (default: {False}).
        '''

        if not is_overwrite and os.path.isfile(filename):
            logger.error('The destination file {} is already exist, abort!'.format(filename))
            sys.exit(1)

        if os.path.isfile(filename):
            wb = load_workbook(filename)
            sheets = wb.get_sheet_names()
            if sheet and sheet in sheets:
                ws = wb.get_sheet_by_name(sheet)
                wb.remove_sheet(ws)

            ws = wb.create_sheet(sheet, index=sheet_index)
            if not sheet:
                ws = wb.active
        else:
            wb = Workbook()
            if sheet:
                ws = wb.create_sheet(sheet, index=sheet_index)
            else:
                ws = wb.active

        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = True
        wb._named_styles = NamedStyleList()

        max_width = {}
        for row_id, row in enumerate(data, start=1):
            for column_id, column in enumerate(row, start=1):
                # for column width
                last_width = max_width[column_id] if column_id in max_width else 0
                # calcalated the column
                unicode_column = column if isinstance(column, unicode) else str(column).decode('utf-8')
                chinese_chars = chinese_counter(unicode_column)
                curr_width = len(unicode_column) + chinese_chars + 2
                # update the biggest column width
                if curr_width > last_width:
                    chinese_chars = chinese_counter(unicode_column)
                    # chinese has 2 characters
                    max_width[column_id] = curr_width

                # for table header
                if has_title and row_id == 1:
                    ws.cell(row=1, column=column_id, value=column).style = self.__header_style
                else:
                    # write data
                    cell = ws.cell(row=row_id, column=column_id, value=column)
                    cell.border = self.__border
                    cell.alignment = self.__alignment

        # set column width
        for k, v in max_width.items():
            ws.column_dimensions[get_column_letter(k)].width = v
        # save file
        wb.save(filename)

    def read(self, file, sheet, rows=[], columns=[]):
        '''Read data from excel file.

        Arguments:
            file {str} -- The absolute path of the destination file
            sheet {str ot int} -- The name or index of the destination sheet

        Keyword Arguments:
            rows {list} -- Row filters, a list which cantains destination row numbers (default: {[]})
            columns {list} -- Column filters, a list which cantains destination column numbers (default: {[]})

        Returns:
            list -- A list of rows, and a tuple for each row
        '''

        if not os.path.isfile(file):
            logger.error('The destination file {} does\'t not exist, abort.'.format(file))
            sys.exit(1)
        if rows and not isinstance(rows, list):
            logger.error('Type error, parameter rows must be a list.')
            sys.exit(1)
        if columns and not isinstance(columns, list):
            logger.error('Type error, parameter columns must be a list.')
            sys.exit(1)
        if rows and filter(lambda x: not isinstance(x, int) or x <= 0, rows):
            logger.error('The row filters must be non zero positive integers.')
            sys.exit(1)
        if columns and filter(lambda x: not isinstance(x, int) or x <= 0, columns):
            logger.error('The column filters must be non zero positive integers.')
            # sys.exit(1)
        try:
            wb = load_workbook(file)
            if isinstance(sheet, str):
                ws = wb.get_sheet_by_name(sheet)
            elif isinstance(sheet, int):
                ws = wb.get_sheet_by_name(wb.get_sheet_names()[sheet - 1])
            else:
                logger.error(
                    'Invalid value for parameter sheet, must be string for sheet name or int for sheet index.'
                )
                sys.exit(1)

            dest_row_index = rows if rows else [i + 1 for i in range(ws.max_row + 1)]
            dest_col_index = columns if columns else [i + 1 for i in range(ws.max_column + 1)]
            dest_data = [
                tuple([ws.cell(row=x, column=y).value for y in dest_col_index]) for x in dest_row_index
            ]
            return dest_data
        except BadZipfile:
            logger.error('Load excel file {} failed, invalid file format.'.format(file))
            sys.exit(1)
        except KeyError:
            logger.error('Worksheet {} does\'t exist.'.format(sheet))
            sys.exit(1)


if __name__ == "__main__":
    excel = ExcelUtil()
    # Usage for write function
    # logger.info('Startting write...')
    # title = (('ID', u'中文', 'DESCRIPTION'))
    # data = [(0, u'苹果', u'手机和电脑'), (1, u'微软', 'Opertion System'), (2, 'Lenove', '')]
    # data = {'title': title, 'data': data}
    # excel.write(filename='c:/Users/Shy/Desktop/w_test.xlsx', data=data, sheet=u'测试', is_overwrite=True)
    # logger.info('Wrtie success!')

    dest_file = '/mnt/c/Users/shy/OneDrive/migu/对外供数/分省/分省供数账户信息.xlsx'
    ip_list = '/home/shy/projects/ip.list'
    source = {}
    target = {}
    # # Usage for read function
    logger.info('Startting read...')
    data = excel.read(dest_file, sheet=1, columns=[3, 5])
    for row in data:
        provence_name = row[0]
        if (row[1] is None):
            continue
        ips = row[1]
        source[provence_name] = ips
    # print(source)
    logger.info('Read finish!')

    for k, v in source.items():
        target[k] = set()

    not_founds = []
    with open(ip_list, 'r') as f:
        for ip in f:
            ip = ip.strip()
            print('Lookup for ip: {}'.format(ip))

            is_found = False
            for k, v in source.items():
                print('Looking in provence: {} with ips: {}...'.format(k, v), end=', ')
                if ip in v:
                    is_found = True
                    target[k].add(ip)
                    print('found!')

            if not is_found:
                print('not found!')
                not_founds.append(ip)
    print(target)

    for k, v in target.items():
        print('{}\t{}'.format(k, ';'.join(list(v))))

    for i in not_founds:
        print(i)
